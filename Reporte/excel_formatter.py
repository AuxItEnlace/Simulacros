from typing import List, Optional
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

HEADER_DARK = "002060"
TAG_COMP = "4f81bd"
TAG_EVID = "f79646"
DATA_WHITE = "FFFFFF"
DATA_COMP = "b8cce4"
DATA_EVID = "fde9d9"
AVG_FILL = "daeef3"
SEP_BLACK = "000000"
PROM_TITLE = "4bacc6"

# Border
_thin = Side(style="thin")
_THIN_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# Fills
_fill_dark = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
_fill_comp = PatternFill(start_color=TAG_COMP, end_color=TAG_COMP, fill_type="solid")
_fill_evid = PatternFill(start_color=TAG_EVID, end_color=TAG_EVID, fill_type="solid")
_fill_white = PatternFill(start_color=DATA_WHITE, end_color=DATA_WHITE, fill_type="solid")
_fill_data_comp = PatternFill(start_color=DATA_COMP, end_color=DATA_COMP, fill_type="solid")
_fill_data_evid = PatternFill(start_color=DATA_EVID, end_color=DATA_EVID, fill_type="solid")
_fill_avg = PatternFill(start_color=AVG_FILL, end_color=AVG_FILL, fill_type="solid")
_fill_black = PatternFill(start_color=SEP_BLACK, end_color=SEP_BLACK, fill_type="solid")
_fill_prom = PatternFill(start_color=PROM_TITLE, end_color=PROM_TITLE, fill_type="solid")

_font_header = Font(bold=True, color="FFFFFF")
_font_tag = Font(bold=True, color="FFFFFF")
_font_std = Font(bold=True, color="FFFFFF")
_align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _border_range(ws: Worksheet, row: int, c1: int, c2: int):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).border = _THIN_BORDER


def _fill_range(ws: Worksheet, row: int, c1: int, c2: int, fill: PatternFill):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill


def _style_range(ws: Worksheet, row: int, c1: int, c2: int, fill: PatternFill,
                 font: Optional[Font] = None, align: Optional[Alignment] = None):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = _THIN_BORDER
        if font:
            cell.font = font
        if align:
            cell.alignment = align


def _merge_write(ws: Worksheet, r1: int, c1: int, r2: int, c2: int,
                 value: str, fill: PatternFill, font: Optional[Font] = None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    cell.value = value
    cell.fill = fill
    cell.font = font or _font_header
    cell.alignment = _align_center
    cell.border = _THIN_BORDER


def _auto_fit_columns(ws: Worksheet):
    """Set column widths to fit the longest value in each column."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col_cells:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            try:
                val = cell.value
            except AttributeError:
                continue  # MergedCell – no direct value
            if val is not None:
                length = len(str(val))
                if length > max_len:
                    max_len = length
        if col_letter and max_len > 0:
            ws.column_dimensions[col_letter].width = min(max_len * 1.15 + 2, 60)


# ---------------------------------------------------------------------------
# Base sheets  (Reporte General, Detallado Grados)
# Two‑row header: static cols merged V‑wise, comp/evid tags in row 1,
# individual names in row 2.
# ---------------------------------------------------------------------------

def format_base_sheet(
    ws: Worksheet,
    n_static: int,
    comp_start: int,
    comp_end: int,
    evid_start: int,
    evid_end: int,
):
    """Insert row 1 and build a two‑row coloured header."""
    max_col = evid_end

    # -- Insert blank row 1 to make room for tag row --
    ws.insert_rows(1, 1)
    # Now: row 2 = old header (individual names), row 3+ = data

    # -- Row 1: static values lifted from row 2 + COMPETENCIAS / EVIDENCIAS --
    for c in range(1, n_static + 1):
        val = ws.cell(row=2, column=c).value  # original header value
        # Clear the bottom cell content first, before it becomes a MergedCell
        ws.cell(row=2, column=c).value = None
        _merge_write(ws, 1, c, 2, c, val or "", _fill_dark)

    if comp_end >= comp_start:
        _merge_write(ws, 1, comp_start, 1, comp_end, "COMPETENCIAS", _fill_comp, _font_tag)
    if evid_end >= evid_start:
        _merge_write(ws, 1, evid_start, 1, evid_end, "EVIDENCIAS", _fill_evid, _font_tag)

    # -- Row 2: individual headers for comp + evid --
    for c in range(comp_start, evid_end + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = _fill_dark
        cell.font = _font_header
        cell.alignment = _align_center
        cell.border = _THIN_BORDER

    # -- Set 27px height for both header rows --
    ws.row_dimensions[1].height = 27
    ws.row_dimensions[2].height = 27

    # -- Data rows (3 … max_row) --
    data_start = 3
    data_end = ws.max_row
    for r in range(data_start, data_end + 1):
        _border_range(ws, r, 1, max_col)
        _fill_range(ws, r, 1, n_static, _fill_white)
        _fill_range(ws, r, comp_start, comp_end, _fill_data_comp)
        _fill_range(ws, r, evid_start, evid_end, _fill_data_evid)


# ---------------------------------------------------------------------------
# School std-dev row  (Reporte General last row)
# ---------------------------------------------------------------------------

def format_school_std_row(ws: Worksheet):
    """Last row: title (col A) as header, value (col B) as data, rest unformatted."""
    row = ws.max_row
    max_col = ws.max_column
    ws.row_dimensions[row].height = 27

    # Clear any formatting left by format_base_sheet on cols C..max_col
    for c in range(3, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()
        cell.font = Font()
        cell.alignment = Alignment()

    # Column A – header style (dark, white bold)
    ca = ws.cell(row=row, column=1)
    ca.fill = _fill_dark
    ca.font = _font_header
    ca.alignment = _align_center
    ca.border = _THIN_BORDER

    # Column B – data style (thin border, centered)
    cb = ws.cell(row=row, column=2)
    cb.border = _THIN_BORDER
    cb.alignment = _align_center


# ---------------------------------------------------------------------------
# BD-RESULTADOS  (single header row, no double-row)
# ---------------------------------------------------------------------------

def format_bd_sheet(ws: Worksheet):
    """Header row dark (30 px), data rows with thin border."""
    max_col = ws.max_column
    ws.row_dimensions[1].height = 30
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _fill_dark
        cell.font = _font_header
        cell.alignment = _align_center
        cell.border = _THIN_BORDER
    for r in range(2, ws.max_row + 1):
        _border_range(ws, r, 1, max_col)


# ---------------------------------------------------------------------------
# Grade sheets  GRADO (0 … 11)
# ---------------------------------------------------------------------------

def format_grade_sheet(ws: Worksheet):
    """Apply two‑row header, data‑row banding, and summary formatting.

    Column layout (same as Detallado Grados):
        A–F static, G–L COMPETENCIAS, M–AJ EVIDENCIAS
    """
    max_col = ws.max_column
    _font_black = Font(color="000000")
    _no_border = Border()

    # ------------------------------------------------------------------
    # 1. Two‑row header
    #
    # Grade‑sheet column layout (0‑based DataFrame index):
    #   A–J (1‑10)  CONTAR, PERCENTIL, PUNTAJE, PUESTO, SEDE, NOMBRE,
    #               GRADO, TOTAL PREGUNTAS, CORRECTAS, % CORRECTAS  → static
    #   K–P (11‑16) Correctas C1 … % Correctas C3                   → COMPETENCIAS
    #   Q+  (17‑…)  Correctas EI1 … % Correctas EI12                → EVIDENCIAS
    # ------------------------------------------------------------------
    n_static = 10
    comp_start = 11
    comp_end = 16
    evid_start = 17
    evid_end = max_col

    ws.insert_rows(1, 1)
    for c in range(1, n_static + 1):
        val = ws.cell(row=2, column=c).value
        ws.cell(row=2, column=c).value = None
        _merge_write(ws, 1, c, 2, c, val or "", _fill_dark)

    if comp_end >= comp_start:
        _merge_write(ws, 1, comp_start, 1, comp_end, "COMPETENCIAS", _fill_comp, _font_tag)
    if evid_end >= evid_start:
        _merge_write(ws, 1, evid_start, 1, evid_end, "EVIDENCIAS", _fill_evid, _font_tag)

    for c in range(comp_start, evid_end + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = _fill_dark
        cell.font = _font_header
        cell.alignment = _align_center
        cell.border = _THIN_BORDER

    ws.row_dimensions[1].height = 27
    ws.row_dimensions[2].height = 27

    # ------------------------------------------------------------------
    # 2. Locate summary rows (header_avg / prom / separators)
    #    The summary labels live in column G (7), regardless of insert.
    # ------------------------------------------------------------------
    prom_rows: List[int] = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=7).value
        if val is not None and isinstance(val, str) and "PROMEDIO" in val:
            prom_rows.append(r)

    summary_set: set = set()
    for pr in prom_rows:
        summary_set.add(pr)
        summary_set.add(pr - 1)                               # header_avg
        if pr + 1 <= ws.max_row:
            summary_set.add(pr + 1)                           # white sep
        if pr + 2 <= ws.max_row:
            summary_set.add(pr + 2)                           # black sep

    # ------------------------------------------------------------------
    # 3. Data‑row banding (white static / comp / evid fills + borders)
    # ------------------------------------------------------------------
    for r in range(3, ws.max_row + 1):
        if r in summary_set:
            continue
        _border_range(ws, r, 1, evid_end)
        _fill_range(ws, r, 1, n_static, _fill_white)
        _fill_range(ws, r, comp_start, comp_end, _fill_data_comp)
        _fill_range(ws, r, evid_start, evid_end, _fill_data_evid)

    # ------------------------------------------------------------------
    # 4. Summary‑row formatting
    # ------------------------------------------------------------------
    for prom_row in reversed(prom_rows):
        head_row = prom_row - 1
        white_row = prom_row + 1
        black_row = prom_row + 2

        # --- Strip residual banding from summary rows ---
        for r in (head_row, prom_row, white_row, black_row):
            if r > ws.max_row:
                continue
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                try:
                    cell.fill = PatternFill(fill_type=None)
                except AttributeError:
                    pass
                try:
                    cell.border = _no_border
                except AttributeError:
                    pass
                try:
                    cell.font = Font()
                except AttributeError:
                    pass
                try:
                    cell.alignment = Alignment()
                except AttributeError:
                    pass

        # --- Separator rows ---
        if white_row <= ws.max_row:
            ws.merge_cells(start_row=white_row, start_column=1, end_row=white_row, end_column=max_col)
            ws.row_dimensions[white_row].height = 13
            for c in range(1, max_col + 1):
                cell = ws.cell(row=white_row, column=c)
                cell.fill = _fill_white
                cell.border = _no_border

        if black_row <= ws.max_row:
            ws.merge_cells(start_row=black_row, start_column=1, end_row=black_row, end_column=max_col)
            ws.row_dimensions[black_row].height = 6
            for c in range(1, max_col + 1):
                cell = ws.cell(row=black_row, column=c)
                cell.fill = _fill_black
                cell.border = _no_border

        # --- header_avg row ---
        # A–F: merged blank, no border
        ws.merge_cells(start_row=head_row, start_column=1, end_row=head_row, end_column=6)
        cell = ws.cell(row=head_row, column=1)
        cell.fill = _fill_white
        cell.border = _no_border
        cell.alignment = _align_center

        # G: "TOTAL ESTUDIANTES" – header style
        cell = ws.cell(row=head_row, column=7)
        cell.fill = _fill_dark
        cell.font = _font_header
        cell.alignment = _align_center
        cell.border = _THIN_BORDER

        # H: value – data style
        cell = ws.cell(row=head_row, column=8)
        cell.fill = _fill_data_comp
        cell.font = _font_black
        cell.alignment = _align_center
        cell.border = _THIN_BORDER

        # I: "DESVIACION ESTANDAR" – header style
        cell = ws.cell(row=head_row, column=9)
        cell.fill = _fill_dark
        cell.font = _font_header
        cell.alignment = _align_center
        cell.border = _THIN_BORDER

        # J: value – data style
        cell = ws.cell(row=head_row, column=10)
        cell.fill = _fill_data_comp
        cell.font = _font_black
        cell.alignment = _align_center
        cell.border = _THIN_BORDER

        # K+ : data style
        for c in range(11, max_col + 1):
            cell = ws.cell(row=head_row, column=c)
            cell.border = _THIN_BORDER
            cell.fill = _fill_data_comp
            cell.font = _font_black
            cell.alignment = _align_center

        # --- Promedio row ---
        # A–F: merged blank, no border
        ws.merge_cells(start_row=prom_row, start_column=1, end_row=prom_row, end_column=6)
        cell = ws.cell(row=prom_row, column=1)
        cell.fill = _fill_white
        cell.border = _no_border
        cell.alignment = _align_center

        # G–H: merged for "PROMEDIO" label (#4bacc6)
        ws.merge_cells(start_row=prom_row, start_column=7, end_row=prom_row, end_column=8)
        cell = ws.cell(row=prom_row, column=7)
        cell.fill = _fill_prom
        cell.font = Font(bold=True, color="000000")
        cell.alignment = _align_center
        cell.border = _THIN_BORDER
        for cc in range(1, max_col + 1):
            v = ws.cell(row=prom_row, column=cc).value
            if v is not None and isinstance(v, str) and "PROMEDIO" in v:
                cell.value = v
                if cc != 7:
                    ws.cell(row=prom_row, column=cc).value = None
                break

        # I–J: J data on both rows; I data on prom_row only (head_row already header)
        for c in (9, 10):
            for r in (head_row, prom_row):
                cell = ws.cell(row=r, column=c)
                if r == head_row and c == 9:
                    continue
                cell.fill = _fill_data_comp
                cell.font = _font_black
                cell.alignment = _align_center
                cell.border = _THIN_BORDER

        # K+ : merged (prom values move up, #daeef3)
        for c in range(11, max_col + 1):
            prom_val = ws.cell(row=prom_row, column=c).value
            if prom_val is not None and prom_val != "":
                ws.cell(row=head_row, column=c).value = prom_val
                ws.cell(row=prom_row, column=c).value = None

            ws.cell(row=head_row, column=c).border = _THIN_BORDER
            ws.cell(row=prom_row, column=c).border = _THIN_BORDER

            ws.merge_cells(
                start_row=head_row, start_column=c,
                end_row=prom_row, end_column=c,
            )
            cell = ws.cell(row=head_row, column=c)
            cell.fill = _fill_avg
            cell.font = _font_black
            cell.alignment = _align_center

        # --- Row heights ---
        ws.row_dimensions[head_row].height = 30
        ws.row_dimensions[prom_row].height = 30


# ---------------------------------------------------------------------------
# Detallado Grados – summary rows  (Desviación Estándar / Promedio)
# ---------------------------------------------------------------------------

def format_detail_summary(ws: Worksheet):
    """Style per‑grade summary rows (Desviación Estándar + Promedio + separators).

    Must be called *after* ``format_base_sheet`` so the header row is already
    inserted and data‑banding has been applied.
    Separator rows must already exist (written by ``add_summary_by_group``).
    """
    max_col = ws.max_column
    _font_black = Font(color="000000")

    # Locate every Desviación Estándar row (col C)
    desv_rows: List[int] = []
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=3).value
        if val is not None and isinstance(val, str) and "Desviación Estándar" in val:
            desv_rows.append(r)

    for desv_row in reversed(desv_rows):
        prom_row = desv_row + 1
        sep_row1 = prom_row + 1  # white separator (written as empty row)
        sep_row2 = prom_row + 2  # black separator

        # --- Strip banding left by format_base_sheet ---
        for r in (desv_row, prom_row, sep_row1, sep_row2):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.font = Font()
                cell.alignment = Alignment()

        # --- A–B: single 2×2 merged cell, white fill ---
        ws.merge_cells(
            start_row=desv_row, start_column=1,
            end_row=prom_row, end_column=2,
        )
        cell = ws.cell(row=desv_row, column=1)
        cell.fill = _fill_white
        cell.border = _THIN_BORDER
        cell.alignment = _align_center

        # --- C: Desviación Estándar title (dark header) ---
        cell = ws.cell(row=desv_row, column=3)
        cell.fill = _fill_dark
        cell.font = _font_header
        cell.alignment = _align_center
        cell.border = _THIN_BORDER

        # --- C: Promedio title (#4bacc6, white bold) ---
        cell = ws.cell(row=prom_row, column=3)
        cell.fill = _fill_prom
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = _align_center
        cell.border = _THIN_BORDER

        # --- D: both rows with #b8cce4, black text, border ---
        for r in (desv_row, prom_row):
            cell = ws.cell(row=r, column=4)
            cell.fill = _fill_data_comp
            cell.font = _font_black
            cell.alignment = _align_center
            cell.border = _THIN_BORDER

        # --- E … max_col: move prom values up, merge, #daeef3 ---
        for c in range(5, max_col + 1):
            prom_val = ws.cell(row=prom_row, column=c).value
            if prom_val is not None and prom_val != "":
                ws.cell(row=desv_row, column=c).value = prom_val
                ws.cell(row=prom_row, column=c).value = None

            # Apply thin border to BOTH cells before merging
            ws.cell(row=desv_row, column=c).border = _THIN_BORDER
            ws.cell(row=prom_row, column=c).border = _THIN_BORDER

            ws.merge_cells(
                start_row=desv_row, start_column=c,
                end_row=prom_row, end_column=c,
            )
            cell = ws.cell(row=desv_row, column=c)
            cell.fill = _fill_avg
            cell.font = _font_black
            cell.alignment = _align_center

        # --- Row heights ---
        ws.row_dimensions[desv_row].height = 30
        ws.row_dimensions[prom_row].height = 30

        # --- Separator rows (already present in the sheet) ---
        for sr, fill, h in ((sep_row1, _fill_white, 13), (sep_row2, _fill_black, 6)):
            ws.merge_cells(
                start_row=sr, start_column=1,
                end_row=sr, end_column=max_col,
            )
            ws.row_dimensions[sr].height = h
            for c in range(1, max_col + 1):
                cell = ws.cell(row=sr, column=c)
                cell.fill = fill
                cell.border = Border()
                cell.font = Font()
                cell.alignment = Alignment()


# ---------------------------------------------------------------------------
# Column ranges
# ---------------------------------------------------------------------------

# Reporte General:   A–E static, F–K COMP, L–AI EVID
STATIC_BASE = 5
COMP_BASE_S = 6
COMP_BASE_E = 11
EVID_BASE_S = 12
EVID_BASE_E = 35

# Detallado Grados:  A–F static, G–L COMP, M–AJ EVID
STATIC_DETAIL = 6
COMP_DETAIL_S = 7
COMP_DETAIL_E = 12
EVID_DETAIL_S = 13
EVID_DETAIL_E = 36


def format_workbook(ws_report, ws_detail, ws_bd, grade_sheets):
    """Apply all formatting in correct order."""
    format_base_sheet(ws_report, STATIC_BASE, COMP_BASE_S, COMP_BASE_E, EVID_BASE_S, EVID_BASE_E)
    format_school_std_row(ws_report)

    format_base_sheet(ws_detail, STATIC_DETAIL, COMP_DETAIL_S, COMP_DETAIL_E, EVID_DETAIL_S, EVID_DETAIL_E)
    format_detail_summary(ws_detail)

    format_bd_sheet(ws_bd)

    for ws in grade_sheets:
        format_grade_sheet(ws)

    # Auto‑fit columns on every sheet
    _auto_fit_columns(ws_report)
    _auto_fit_columns(ws_detail)
    _auto_fit_columns(ws_bd)
    for ws in grade_sheets:
        _auto_fit_columns(ws)
