import statistics
import re
from pathlib import Path
from typing import Dict, List, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

C_COUNT = 3
EI_COUNT = 12
GRADE_RANGE = 12

INT_AVG_COLS: List[str] = (
    ["CORRECTAS"]
    + [f"Correctas C{i}" for i in range(1, C_COUNT + 1)]
    + [f"Correctas EI{i}" for i in range(1, EI_COUNT + 1)]
)

PCT_AVG_COLS: List[str] = (
    ["% CORRECTAS"]
    + [f"% Correctas C{i}" for i in range(1, C_COUNT + 1)]
    + [f"% Correctas EI{i}" for i in range(1, EI_COUNT + 1)]
)

AVG_COLS: List[str] = [
    x for pair in zip(INT_AVG_COLS, PCT_AVG_COLS) for x in pair
]

# Columns to remove when building per-grade sheets
DETAIL_COLUMNS_TO_DROP: List[str] = (
    [str(i) for i in range(1, 43)]
    + ["DOCUMENTO"]
    + [f"Explicación {i}" for i in range(1, 43)]
    + [
        "INCORRECTAS",
        "% INCORRECTAS",
        "INVÁLIDAS",
        "% INVÁLIDAS",
        "VACÍAS",
        "% VACÍAS",
    ]
)


# ---------------------------------------------------------------------------
# Sheet creation
# ---------------------------------------------------------------------------

def create_base_sheets(wb: Workbook, sheet_names: List[str]) -> Workbook:
    ws = wb.active
    ws.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(name)
    return wb


# ---------------------------------------------------------------------------
# Headers
# ---------------------------------------------------------------------------

def _competence_headers() -> List[str]:
    headers: List[str] = []
    for i in range(1, 4):
        headers.append(f"Correctas C{i}")
        headers.append(f"% Correctas C{i}")
    return headers


def _ei_headers() -> List[str]:
    headers: List[str] = []
    for i in range(1, 13):
        headers.append(f"Correctas EI{i}")
        headers.append(f"% Correctas EI{i}")
    return headers


def build_base_header() -> List[str]:
    return (
        ["COLEGIO", "GRADO", "DESVESTA", "CORRECTAS", "%CORRECTAS"]
        + _competence_headers()
        + _ei_headers()
    )


def build_detailed_header() -> List[str]:
    return (
        ["PERCENTIL", "PUESTO", "GRADO", "DESVESTA", "CORRECTAS", "%CORRECTAS"]
        + _competence_headers()
        + _ei_headers()
    )


# ---------------------------------------------------------------------------
# Write raw data into a sheet
# ---------------------------------------------------------------------------

def write_dataframe_to_sheet(wb: Workbook, sheet_name: str, data: pd.DataFrame) -> None:
    for row in dataframe_to_rows(data, index=False, header=True):
        wb[sheet_name].append(row)


# ---------------------------------------------------------------------------
# Per-grade sheets (average calculation per group within each grade)
# ---------------------------------------------------------------------------

def process_grade_sheets(
    wb: Workbook, data: pd.DataFrame, grade_sheets: List[str]
) -> List[List[Any]]:
    """Fill each grade sheet with student rows + average rows.

    Returns a list of summary rows, each with the structure::

        [grado_str, std_dev, CORRECTAS_avg, %CORRECTAS_avg_str,
         CorrectasC1_avg, %CorrectasC1_avg_str, …]
    """
    avg_data: List[List[Any]] = []
    data = data.copy()
    data["GRADO_NUM"] = data["GRADO"].str.extract(r"^(\d+)").astype(int)
    max_col_grade: Optional[int] = None

    for sheet_name in grade_sheets:
        grade_num = int(sheet_name.split("(")[1].split(")")[0])
        grade_data = (
            data[data["GRADO_NUM"] == grade_num].copy()
        )
        grade_data.drop(columns=["GRADO_NUM"], inplace=True)
        grade_data.drop(columns=DETAIL_COLUMNS_TO_DROP, inplace=True, errors="ignore")

        sheet = wb[sheet_name]
        sheet.append(grade_data.columns.tolist())
        n_cols = len(grade_data.columns)
        if max_col_grade is None:
            max_col_grade = n_cols

        groups = list(grade_data.groupby("GRADO"))
        for idx, (grado, group) in enumerate(groups):
            for row in group.itertuples(index=False):
                sheet.append(list(row))

            valid = group[group["CORRECTAS"] > 0]

            pct_series = (
                valid["% CORRECTAS"]
                .astype(str)
                .str.replace("%", "")
                .astype(float)
            )
            std_dev = round(pct_series.std(), 2)

            # -- average row (ordered by DataFrame columns) --
            avg_row: List[Any] = []
            for col in grade_data.columns:
                if col in AVG_COLS:
                    if col in INT_AVG_COLS:
                        avg_row.append(
                            round(valid[col].astype(float).mean(), 2)
                        )
                    else:
                        m = round(
                            valid[col]
                            .astype(str)
                            .str.replace("%", "")
                            .astype(float)
                            .mean(),
                            2,
                        )
                        avg_row.append(f"{m}%")
                elif col == "GRADO":
                    avg_row.append(f"PROMEDIO {grado}")
                else:
                    avg_row.append("")

            header_avg: List[Any] = [
                "",
                "",
                "",
                "",
                "",
                "",
                "TOTAL ESTUDIANTES",
                len(valid),
                "DESVIACION ESTANDAR",
                std_dev,
            ]
            # Pad header_avg to match column count
            while len(header_avg) < n_cols:
                header_avg.append("")

            sheet.append(header_avg)
            sheet.append(avg_row)

            # -- separator rows AFTER each group --
            sheet.append([""] * n_cols)          # white (13px)
            sheet.append([""] * n_cols)          # black (6px)

            # -- collect summary row (ordered by AVG_COLS) --
            summary: List[Any] = [grado, std_dev]
            for col in AVG_COLS:
                if col in INT_AVG_COLS:
                    summary.append(round(valid[col].astype(float).mean(), 2))
                else:
                    m = round(
                        valid[col]
                        .astype(str)
                        .str.replace("%", "")
                        .astype(float)
                        .mean(),
                        2,
                    )
                    summary.append(f"{m}%")
            avg_data.append(summary)

    return avg_data


# ---------------------------------------------------------------------------
# Summary by grade group (percentiles, ranks, averages)
# ---------------------------------------------------------------------------

def _parse_numeric(val: Any) -> float:
    """Try to convert *val* to float, returning 0.0 on failure."""
    if isinstance(val, str):
        val = val.replace("%", "").replace(",", ".")
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def add_summary_by_group(
    ws, avg_sorted: List[List[Any]], school_name: str
) -> List[List[Any]]:
    """Write grouped summary rows (percentile, rank, averages) into *ws*.

    Returns a list of per-grade-group averages ready for the base report sheet.
    """
    # Group rows by grade number
    groups: Dict[int, List[List[Any]]] = {}
    for row in avg_sorted:
        grado: str = str(row[0])
        m = re.match(r"(\d+)", grado)
        if m is None:
            continue
        numero = int(m.group(1))
        groups.setdefault(numero, []).append(list(row))

    grade_averages: List[List[Any]] = []

    for numero in sorted(groups):
        filas = groups[numero]

        # Puntajes for percentile / rank – use the % CORRECTAS column (index 3)
        puntajes = [_parse_numeric(f[3]) for f in filas]

        percentiles = _calc_percentiles(puntajes)
        puestos = [sorted(puntajes, reverse=True).index(p) + 1 for p in puntajes]

        modified: List[List[Any]] = []
        for i, fila in enumerate(filas):
            # Row average (all numeric columns except grade name)
            numeric_vals = [_parse_numeric(v) for v in fila[1:]]
            row_avg = (
                round(sum(numeric_vals) / len(numeric_vals), 2)
                if numeric_vals
                else 0
            )
            new_row = [percentiles[i], puestos[i]] + fila
            modified.append(new_row)
        modified.sort(key=lambda x: x[1])  # sort by rank

        # Write rows
        for row in modified:
            ws.append(row)

        # Find the first empty column index (fallback = 2)
        first_row = modified[0]
        try:
            idx_empty = first_row.index("")
        except ValueError:
            idx_empty = 2

        # Std dev of % CORRECTAS (now at col idx_empty + 2 because of the
        # two prepended columns PERCENTIL / PUESTO)
        col_correctas = idx_empty + 2
        correctas_vals = [_parse_numeric(f[col_correctas]) for f in modified]
        std = (
            round(statistics.stdev(correctas_vals), 2)
            if len(correctas_vals) > 1
            else 0
        )

        # Std dev row
        std_row: List[Any] = [""] * (idx_empty + 3)
        std_row[2] = "Desviación Estándar"
        std_row[idx_empty + 1] = std
        ws.append(std_row)

        # Column averages (from idx_empty onward)
        num_cols = len(modified[0])
        col_avgs: List[Any] = []
        for ci in range(idx_empty, num_cols):
            vals = [_parse_numeric(f[ci]) for f in modified]
            avg = round(sum(vals) / len(vals), 2) if vals else 0
            if (
                isinstance(modified[0][ci], str)
                and modified[0][ci].endswith("%")
            ):
                avg = f"{avg}%"
            col_avgs.append(avg)

        avg_row: List[Any] = [""] * idx_empty
        avg_row.extend(col_avgs)
        avg_row[2] = "Promedio"
        ws.append(avg_row)

        # Separator rows (white + black) – formatted later
        ncols = len(modified[0])
        ws.append([""] * ncols)
        ws.append([""] * ncols)

        grade_averages.append(
            [school_name, f"{numero}°", std] + col_avgs[2:]
        )

    return grade_averages


def _calc_percentiles(scores: List[float]) -> List[int]:
    n = len(scores)
    sorted_scores = sorted(scores)
    result: List[int] = []
    for p in scores:
        positions = [i + 1 for i, v in enumerate(sorted_scores) if v == p]
        avg_rank = sum(positions) / len(positions)
        result.append(round((avg_rank / n) * 100))
    return result


# ---------------------------------------------------------------------------
# Format percentages in Excel cells
# ---------------------------------------------------------------------------

def format_percentages(wb: Workbook, sheet_names: List[str]) -> None:
    for name in sheet_names:
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
            for cell in row:
                try:
                    val = cell.value
                except AttributeError:
                    continue  # MergedCell – no direct value
                if isinstance(val, str) and val.strip().endswith("%"):
                    try:
                        cell.value = float(val.strip().replace("%", "")) / 100
                        cell.number_format = "0%"
                    except ValueError:
                        pass
                elif isinstance(val, (int, float)):
                    cell.number_format = "0" if val == int(val) else "0.0"


# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------

def save_workbook(wb: Workbook, config: Dict[str, str], output_dir: str = "data") -> str:
    sede = config.get("Sede", "")
    school_part = f'{config["SchoolName"]} {sede}' if sede else config["SchoolName"]
    filename = f'{school_part} {config["Date"]} PRUEBA {config["Prueba"]}.xlsx'
    path = Path(output_dir) / filename
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)
